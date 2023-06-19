import os
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


def newBookingFormExport():
    workbook = Workbook()
    sheet = workbook.active
    
    
    #variableCellLocator = ["A1", "B1", "C1", "D1","E1","F1","G1","H1","I1","J1"]

    colmumns = ["A", "B", "C", "D","E","F","G","H","I","J"]
    cellReference = []
    for i, colmumn in enumerate(colmumns):
        cell = colmumn + str(1)  # Append the number to the value
        cellReference.append(cell)
    
    sheet[cellReference[0]] = "Booking ref. "
    sheet[cellReference[1]] = "Booking date"
    sheet[cellReference[2]] = "Lead Guest"
    sheet[cellReference[3]] = "Pax"
    sheet[cellReference[4]] = "Dep' date"
    sheet[cellReference[5]] = "No. nights"
    sheet[cellReference[6]] = "Total value"
    sheet[cellReference[7]] = "Deposit"
    sheet[cellReference[8]] = "Outstanding balance"
    sheet[cellReference[9]] = "Due date"

    current_datetime = datetime.now()
    datetime_str = current_datetime.strftime("%d-%m-%y_%H%M%S")

    file_name = f"{datetime_str}-BookingFormData.xlsx"
    folder_path = ".\excel_exports"

    file_path = os.path.join(folder_path, file_name)
    workbook.save(file_path)

    return file_path

def writeOrderFormData(BookFilePath,LoopItteration,pdfData):
    workbook = load_workbook(BookFilePath)
    sheet = workbook.active
    
    colmumns = ["A", "B", "C", "D","E","F","G","H","I","J"]
    cellReference = []
    for i, colmumn in enumerate(colmumns):
        cell = colmumn + str(LoopItteration)  # Append the number to the value
        cellReference.append(cell)

    sheet[cellReference[0]] = pdfData[0]
    sheet[cellReference[1]] = pdfData[1]
    sheet[cellReference[2]] = pdfData[2]
    sheet[cellReference[3]] = pdfData[3]
    sheet[cellReference[4]] = pdfData[4]
    sheet[cellReference[5]] = pdfData[5]
    sheet[cellReference[6]] = pdfData[6]
    sheet[cellReference[7]] = pdfData[7]
    sheet[cellReference[8]] = pdfData[8]
    sheet[cellReference[9]] = pdfData[9]

    workbook.save(BookFilePath)

    return True